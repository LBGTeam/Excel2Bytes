import os
import pandas as pd
from openpyxl import load_workbook

from LogUtil import ShowLog
from GlobalUtil import LanguageXlsxPath, LanguageDict
from ConfigData import Config


def InitLanguage():
    ShowLog('初始化语言表')
    isUpdateAllLNG = Config.IsUpdateAllLNG()
    cnLanguage = Config.CNLanguage()
    languageKey = Config.LanguageKey()
    if (not os.path.exists(LanguageXlsxPath) or
            cnLanguage not in pd.read_excel(LanguageXlsxPath, sheet_name=None).keys()):
        SaveLangData(cnLanguage, {'c': 'c', 'ID': 'text', 'uint': 'string', '编号': '文本'})
    ReadLangData(cnLanguage)
    if isUpdateAllLNG:
        for key in languageKey:
            ReadLangData(key)


def SaveLanguage():
    ShowLog('保存语言表')
    isUpdateAllLNG = Config.IsUpdateAllLNG()
    cnLanguage = Config.CNLanguage()
    languageKey = Config.LanguageKey()
    SaveLangData(cnLanguage, LanguageDict[cnLanguage])
    if isUpdateAllLNG:
        languageDict = {cnLanguage: {}}
        for key, value in LanguageDict[cnLanguage].items():
            languageDict[cnLanguage][key] = value
            for keyLng in languageKey:
                if keyLng not in languageDict.keys():
                    languageDict[keyLng] = {}
                if key in LanguageDict[keyLng].keys():
                    languageDict[keyLng][key] = LanguageDict[keyLng][key]
                else:
                    languageDict[keyLng][key] = value
        for keyLng in languageKey:
            SaveLangData(keyLng, languageDict[keyLng])


def ReadLangData(key):
    ShowLog(f'读取语言表数据: {key}')
    # 读取Excel文件
    df = pd.read_excel(LanguageXlsxPath, sheet_name=None, header=None)
    # 判断是否包含名为key的页签
    if key in df.keys():
        df = pd.read_excel(LanguageXlsxPath, sheet_name=key)
        keyLng = {}
        for index, row in df.iloc[3:].iterrows():
            keyLng[row.iloc[0]] = row.iloc[1]
        LanguageDict[key] = keyLng


def SaveLangData(key, data):
    ShowLog(f'保存语言表数据: {key}')
    workbook = load_workbook(filename=LanguageXlsxPath)
    if key not in workbook.sheetnames:
        worksheet = workbook.create_sheet(key)
        worksheet.append(['c', 'c'])
        worksheet.append(['ID', 'text'])
        worksheet.append(['uint', 'string'])
        worksheet.append(['编号', '文本'])
    else:
        worksheet = workbook.get_sheet_by_name(key)
        worksheet.delete_rows(5, worksheet.max_row)
    for key, value in data.items():
        worksheet.append([key, value])
    workbook.save(LanguageXlsxPath)


def GetLanguageKey(value):
    if value.isspace() or value == '':
        return 0
    cnLanguage = Config.CNLanguage()
    isUpdateAllLNG = Config.IsUpdateAllLNG()
    languageKey = Config.LanguageKey()
    if value in LanguageDict[cnLanguage].values():
        for key, val in LanguageDict[cnLanguage].items():
            if val == value:
                return key
    else:
        if len(LanguageDict[cnLanguage]) == 0:
            LanguageDict[cnLanguage][1] = value
            return 1
        maxKey = max(LanguageDict[cnLanguage].keys())
        newKey = maxKey + 1
        LanguageDict[cnLanguage][newKey] = value
        if isUpdateAllLNG:
            for key in languageKey:
                LanguageDict[key][newKey] = value
        return newKey

